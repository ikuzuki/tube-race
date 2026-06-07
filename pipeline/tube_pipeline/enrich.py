"""Enrich graph stations with trivia/stats from open data.

For every station already present in ``graph.json`` this module produces a
best-effort :class:`~tube_pipeline.models.StationInfo`: opening year (with a
network rank), an approximate daily traffic figure (with a rank), a short
sourced fun fact, and a link to the Wikipedia article. Every field is optional
-- only what could be resolved from open data is emitted.

Pipeline
--------
1. Query Wikidata (SPARQL) for London Underground stations, returning label,
   coordinates, opening date, and the English Wikipedia article title. (Wikidata
   also carries a patronage field, but its tube coverage is sparse -- see step
   4 -- so it is only a fallback for daily traffic.)
2. Match each Wikidata candidate to a graph station by coordinate proximity
   (haversine, nearest within ~400 m), falling back to a normalised-name
   match.
3. Fetch a one-line fun fact per matched station from the Wikipedia REST
   summary endpoint (polite: descriptive User-Agent, per-request delay).
4. Download TfL's authoritative Annual Station Counts spreadsheet and read each
   Underground station's annualised entries+exits, matched to a graph station
   by normalised name. This covers essentially the whole network, where the
   Wikidata patronage field covers only a handful of stations; the TfL figure
   is therefore preferred and Wikidata patronage is the fallback.
5. Convert the annual figure to a daily one (divide by 365), then rank stations
   by opening year (1 = oldest) and daily traffic (1 = busiest).

Network access is confined to :class:`WikidataClient`, :class:`WikipediaClient`
and :class:`StationUsageClient`; :func:`enrich_stations` takes them as arguments
so it can be driven against mocked HTTP (and an in-memory XLSX) in tests.

See ``SPEC.md`` and ``web/src/lib/stationInfo.ts`` for the output contract.
"""

from __future__ import annotations

import math
import re
import time
import zipfile
from dataclasses import dataclass
from typing import Any

import httpx

from tube_pipeline.curated_facts import CURATED_FACTS
from tube_pipeline.curated_stats import (
    CURATED_DAILY_TRAFFIC,
    CURATED_OPENED_YEARS,
    CURATED_YEAR_CORRECTIONS,
)
from tube_pipeline.models import (
    StationInfo,
    StationInfoCounts,
    StationInfoFile,
)

WIKIDATA_SPARQL_URL: str = "https://query.wikidata.org/sparql"
"""Wikidata SPARQL query-service endpoint."""

WIKIPEDIA_SUMMARY_URL: str = "https://en.wikipedia.org/api/rest_v1/page/summary"
"""Base URL of the Wikipedia REST page-summary endpoint."""

WIKIPEDIA_QUERY_URL: str = "https://en.wikipedia.org/w/api.php"
"""Wikipedia Action API endpoint, used for the fuller ``prop=extracts`` intro.

The REST summary ``extract`` is sometimes a single sentence; the Action API's
``prop=extracts&exintro&explaintext`` returns the whole lead section as plain
text, which gives the fun-fact heuristic more material to choose from.
"""

ANTHROPIC_MESSAGES_URL: str = "https://api.anthropic.com/v1/messages"
"""Anthropic Messages API endpoint (used only when ``ANTHROPIC_API_KEY`` is set)."""

ANTHROPIC_VERSION: str = "2023-06-01"
"""``anthropic-version`` header value for the Messages API."""

ANTHROPIC_MODEL: str = "claude-3-5-haiku-latest"
"""Cheap, fast Haiku model used to distil one grounded fun fact per station."""

ANTHROPIC_MAX_TOKENS: int = 80
"""Token ceiling for the distilled fun fact (one short sentence)."""

ANTHROPIC_TIMEOUT_S: float = 30.0
"""Per-request timeout for the Anthropic call (kept short; falls back on timeout)."""

USER_AGENT: str = (
    "TubeRacePipeline/0.1 (https://github.com/tube-race; station trivia enrichment) httpx"
)
"""Descriptive User-Agent sent to both Wikimedia services (their policy asks for one)."""

DEFAULT_TIMEOUT: float = 60.0
"""Default per-request timeout in seconds (SPARQL can be slow)."""

DEFAULT_MATCH_RADIUS_M: float = 400.0
"""Maximum great-circle distance (metres) for a coordinate match."""

DEFAULT_WIKI_DELAY_S: float = 0.1
"""Polite delay between Wikipedia summary requests (the ~272 per-station calls)."""

SPARQL_MAX_RETRIES: int = 4
"""Attempts for the SPARQL query before giving up on a throttled endpoint."""

SPARQL_MAX_BACKOFF_S: float = 90.0
"""Cap on how long to honour a ``Retry-After`` between SPARQL attempts."""

_RETRYABLE_STATUS: frozenset[int] = frozenset({429, 500, 502, 503, 504})
"""HTTP statuses worth retrying the SPARQL query for (throttling/transient)."""

_EARTH_RADIUS_M: float = 6_371_000.0
"""Mean Earth radius in metres, for the haversine distance."""

STATION_USAGE_URL: str = (
    "https://crowding.data.tfl.gov.uk/"
    "Annual%20Station%20Counts/2023/AC2023_AnnualisedEntryExit.xlsx"
)
"""TfL Annual Station Counts (2023): annualised entries+exits per station.

This is the most recent year published in TfL's open ``crowding.data.tfl.gov.uk``
bucket at time of writing; later years appear under the same path scheme. The
spreadsheet covers London Underground, Overground, DLR and Elizabeth line; only
the Underground (``LU``) rows are used here.
"""

DAYS_PER_YEAR: float = 365.0
"""Divisor turning an annual entries+exits total into an approximate daily one."""

# The Wikidata SPARQL query. A station qualifies if it is served by / has as a
# connecting line (P81) a line that is part of (P361, transitively) the London
# Underground (Q20075). The connecting-line-to-network relationship is the
# distinctive, reliable signal; an explicit station-class constraint is
# deliberately avoided because tube stations are typed inconsistently. The
# qualifying signal is a UNION of three cheap, direct (no property-path closure)
# arms -- a station counts if any of:
#   1. it is a direct instance of a London Underground station (Q14562709);
#   2. its transport network (P16) is the London Underground (Q20075) -- this
#      catches multi-modal interchanges (e.g. Stratford) typed only as railway
#      stations;
#   3. it has a connecting line (P81) whose transport network (P16) is the
#      London Underground (Q20075).
# Direct triples are used throughout: subclass-chain paths over a UNION are
# expensive enough that WDQS rejects the query with HTTP 429. (Earlier dead
# ends came from wrong QIDs and from the line->network link being P16, not
# P361.) Coordinates (P625) are required; opening date (P1619), patronage
# (P3872) and the enwiki article are optional. Note P3872 coverage for tube
# stations is sparse, so daily-traffic output is partial by design.
SPARQL_QUERY: str = """
SELECT DISTINCT ?station ?stationLabel ?coord ?opened ?patronage ?article WHERE {
  { ?station wdt:P31 wd:Q14562709 . }
  UNION
  { ?station wdt:P16 wd:Q20075 . }
  UNION
  { ?station wdt:P81 ?line . ?line wdt:P16 wd:Q20075 . }
  ?station wdt:P625 ?coord .
  OPTIONAL { ?station wdt:P1619 ?opened . }
  OPTIONAL { ?station wdt:P3872 ?patronage . }
  OPTIONAL {
    ?article schema:about ?station ;
             schema:isPartOf <https://en.wikipedia.org/> .
  }
  SERVICE wikibase:label { bd:serviceParam wikibase:language "en". }
}
"""
"""SPARQL used to pull London Underground station candidates from Wikidata."""

_WKT_POINT_RE: re.Pattern[str] = re.compile(r"Point\(\s*(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s*\)")
"""Matches a WKT ``Point(lon lat)`` literal as returned by Wikidata P625."""

_STATION_SUFFIX_RE: re.Pattern[str] = re.compile(
    r"\s*\(.*?\)|\s*(?:underground|tube|dlr|rail|railway)?\s*stations?\b|-underground\b",
    re.IGNORECASE,
)
"""Strips parentheticals and ``... [underground] station`` suffixes for name matching."""

_TITLE_SUFFIX_RE: re.Pattern[str] = re.compile(
    r"\s*\([^)]*\)\s*$"
    r"|[\s-]*(?:underground|tube|rail|dlr)\s*station\s*$"
    r"|-underground\s*$"
    r"|\s+(?:rail|dlr|ell)\s*$",
    re.IGNORECASE,
)
"""End-anchored strip for building a fallback Wikipedia title.

Unlike :data:`_STATION_SUFFIX_RE`, this removes only a trailing parenthetical or
``... [underground] station`` suffix, so an internal ``Station`` (as in
``Battersea Power Station``) is preserved in the article title.
"""


@dataclass(frozen=True)
class WikidataStation:
    """A single London Underground station candidate from Wikidata.

    Parameters
    ----------
    qid : str
        Wikidata entity id, e.g. ``"Q1234"``.
    name : str
        English label.
    lat, lon : float
        WGS84 coordinates from P625.
    opened_year : int or None
        Year parsed from the opening date (P1619), if present.
    annual_patronage : float or None
        Annual passenger usage (P3872), if present.
    article_title : str or None
        English Wikipedia article title (from the sitelink URL), if present.
    """

    qid: str
    name: str
    lat: float
    lon: float
    opened_year: int | None
    annual_patronage: float | None
    article_title: str | None


@dataclass(frozen=True)
class GraphStation:
    """The subset of a graph station needed for enrichment.

    Parameters
    ----------
    id : str
        Graph station id (output key).
    name : str
        Display name from the graph.
    lat, lon : float
        WGS84 coordinates from the graph.
    """

    id: str
    name: str
    lat: float
    lon: float


def haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Great-circle distance between two points in metres.

    Parameters
    ----------
    lat1, lon1 : float
        First point in decimal degrees.
    lat2, lon2 : float
        Second point in decimal degrees.

    Returns
    -------
    float
        Distance in metres.
    """
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2.0) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2.0) ** 2
    return 2.0 * _EARTH_RADIUS_M * math.asin(math.sqrt(a))


def normalise_name(name: str) -> str:
    """Normalise a station name for fallback matching.

    Lowercases, strips ``... station``/``... underground station`` suffixes and
    parentheticals, expands ``&`` to ``and`` and ``st``/``st.`` to ``saint``,
    drops apostrophes and other punctuation, and collapses whitespace.

    Parameters
    ----------
    name : str
        Raw station name from either source.

    Returns
    -------
    str
        A normalised key suitable for equality comparison.
    """
    text = name.lower()
    text = _STATION_SUFFIX_RE.sub(" ", text)
    text = text.replace("&", " and ")
    text = text.replace("'", "").replace("’", "")
    # Expand "st"/"st." as a standalone word to "saint" (St Paul's -> saint pauls).
    text = re.sub(r"\bst\.?\b", "saint", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


_MODE_TAIL_RE: re.Pattern[str] = re.compile(
    r"\b(?:lu|dlr|tfl|elr|nr|lo|ell|ezl)\s*$", re.IGNORECASE
)
"""Trailing mode token TfL appends to disambiguate co-located stations.

In the Annual Station Counts file a station that shares a name with a National
Rail or multi-modal hub carries a ``" LU"`` (or ``" TfL"``) suffix on the
Underground gateline row -- e.g. ``"Euston LU"``, ``"Victoria LU"``,
``"Paddington TfL"`` -- to separate it from the mainline counts. The graph names
have no such suffix, so it is stripped before matching.
"""

_USAGE_NAME_ALIASES: dict[str, tuple[str, ...]] = {
    # One combined Underground row feeds two distinct graph stations.
    "bank and monument": ("bank", "monument"),
    # TfL writes the terminals 2 & 3 row as "Heathrow Terminals 123".
    "heathrow terminals 123": ("heathrow terminals 2 and 3",),
}
"""Maps a normalised usage-file name to the graph key(s) it should populate.

Resolves the handful of cases where TfL's naming differs structurally from the
graph's, rather than merely by the mode suffix handled by :data:`_MODE_TAIL_RE`.
Each value is one or more graph match keys (as produced by
:func:`normalise_name`) the row's figure applies to.
"""


def usage_match_keys(name: str) -> tuple[str, ...]:
    """Map a TfL usage-file station name to graph match key(s).

    Normalises the name the same way graph names are (:func:`normalise_name`),
    strips TfL's trailing mode-disambiguation token (``"... LU"``/``"... TfL"``),
    then expands any structural alias (:data:`_USAGE_NAME_ALIASES`). The result
    is the set of :func:`normalise_name` keys whose graph station this row's
    figure should populate.

    Parameters
    ----------
    name : str
        Raw ``Station`` cell from the Annual Station Counts spreadsheet.

    Returns
    -------
    tuple of str
        Zero or more graph match keys. Empty only when the name normalises away
        to nothing.
    """
    key = _MODE_TAIL_RE.sub("", normalise_name(name)).strip()
    if not key:
        return ()
    return _USAGE_NAME_ALIASES.get(key, (key,))


def parse_wkt_point(literal: str) -> tuple[float, float] | None:
    """Parse a Wikidata ``Point(lon lat)`` WKT literal.

    Parameters
    ----------
    literal : str
        The coordinate value string, e.g. ``"Point(-0.1437 51.4965)"``.

    Returns
    -------
    tuple of (float, float) or None
        ``(lat, lon)`` on success, ``None`` if the literal does not parse.
    """
    match = _WKT_POINT_RE.search(literal)
    if match is None:
        return None
    lon = float(match.group(1))
    lat = float(match.group(2))
    return lat, lon


def parse_opened_year(literal: str) -> int | None:
    """Extract a four-digit year from a Wikidata date literal.

    Wikidata dates are ISO-ish, optionally signed and zero-padded, e.g.
    ``"1868-12-24T00:00:00Z"`` or ``"+1863-01-01T00:00:00Z"``. Only the leading
    year is needed.

    Parameters
    ----------
    literal : str
        The opening-date value string.

    Returns
    -------
    int or None
        The year, or ``None`` if no plausible year is found.
    """
    match = re.match(r"[+-]?(\d{1,4})-", literal)
    if match is None:
        return None
    year = int(match.group(1))
    # Guard against year 0 / obviously bad values; the tube is 19th century on.
    if 1800 <= year <= 2100:
        return year
    return None


def article_title_from_url(url: str) -> str | None:
    """Recover a Wikipedia article title from its sitelink URL.

    Parameters
    ----------
    url : str
        A URL such as ``"https://en.wikipedia.org/wiki/Oval_tube_station"``.

    Returns
    -------
    str or None
        The decoded title (``"Oval tube station"``), or ``None`` if the URL has
        no ``/wiki/`` path segment.
    """
    marker = "/wiki/"
    idx = url.find(marker)
    if idx == -1:
        return None
    from urllib.parse import unquote

    slug = url[idx + len(marker) :]
    if not slug:
        return None
    return unquote(slug).replace("_", " ")


_USAGE_MODE_COLUMN: str = "mode"
"""Header label of the transport-mode column in the Annual Station Counts sheet."""

_USAGE_STATION_COLUMN: str = "station"
"""Header label of the station-name column."""

_USAGE_ANNUAL_HEADER: str = "en/ex"
"""Per-column header of the entries+exits summary columns (weekly/12-week/annual)."""

_USAGE_ANNUAL_GROUP: str = "annualised"
"""Group label (one row above the header) marking the annualised total column."""

_UNDERGROUND_MODE: str = "lu"
"""Value in the mode column identifying a London Underground row."""

_USAGE_MODES: frozenset[str] = frozenset({"lu", "lo", "dlr", "ezl"})
"""Mode-column values the game counts: Underground, Overground, DLR, Elizabeth.

The Annual Station Counts workbook covers every TfL rail mode; we include all of
them so the expanded network (Overground/DLR/Elizabeth stations) gets real
gateline figures rather than estimates. Interchanges that appear under several
modes collapse on the match key, where the larger figure wins.
"""


def _header_row_index(rows: list[tuple[Any, ...]]) -> int | None:
    """Find the column-header row in an Annual Station Counts sheet.

    The sheet has several title rows before the real header; the header is the
    first row carrying both a ``Mode`` and a ``Station`` cell.

    Parameters
    ----------
    rows : list of tuple
        All sheet rows as value tuples.

    Returns
    -------
    int or None
        Zero-based index of the header row, or ``None`` if not found.
    """
    for idx, row in enumerate(rows):
        labels = {str(c).strip().lower() for c in row if c is not None}
        if _USAGE_MODE_COLUMN in labels and _USAGE_STATION_COLUMN in labels:
            return idx
    return None


def _column_index(row: tuple[Any, ...], label: str) -> int | None:
    """Return the index of the first cell in ``row`` matching ``label``.

    Parameters
    ----------
    row : tuple
        A header row.
    label : str
        Lowercased label to find.

    Returns
    -------
    int or None
        Column index, or ``None`` if absent.
    """
    for idx, cell in enumerate(row):
        if cell is not None and str(cell).strip().lower() == label:
            return idx
    return None


def _annualised_column_index(header: tuple[Any, ...], group_row: tuple[Any, ...]) -> int | None:
    """Locate the annualised entries+exits column.

    Three trailing summary columns share the header ``En/Ex`` (weekly, 12-week,
    annualised); they are told apart by the group label one row above. The
    annualised column is the one whose own header is ``En/Ex`` and whose group
    label is ``Annualised`` -- located by label rather than fixed offset so a
    re-ordered or re-sized sheet still parses.

    Parameters
    ----------
    header : tuple
        The column-header row.
    group_row : tuple
        The row immediately above the header (the merged group labels).

    Returns
    -------
    int or None
        Column index of the annualised total, or ``None`` if not found.
    """
    for idx, cell in enumerate(header):
        head = str(cell).strip().lower() if cell is not None else ""
        group = (
            str(group_row[idx]).strip().lower()
            if idx < len(group_row) and group_row[idx] is not None
            else ""
        )
        if head == _USAGE_ANNUAL_HEADER and group == _USAGE_ANNUAL_GROUP:
            return idx
    return None


def parse_station_usage(data: bytes) -> dict[str, float]:
    """Parse annualised entries+exits from the TfL counts XLSX.

    Reads TfL's Annual Station Counts workbook and returns, for every counted
    TfL rail row (Underground, Overground, DLR or Elizabeth; see
    :data:`_USAGE_MODES`) carrying a numeric annualised total, the annual
    entries+exits keyed by graph match key (:func:`usage_match_keys`). The
    header is self-located (see :func:`_header_row_index` and
    :func:`_annualised_column_index`) rather than hardcoded by offset.

    Rows whose annualised cell is non-numeric (TfL writes ``"---"`` where a
    station has no count) are skipped. On a key collision the larger figure
    wins, which keeps the busier of two co-located rows.

    Parameters
    ----------
    data : bytes
        Raw ``.xlsx`` file contents.

    Returns
    -------
    dict of str to float
        Graph match key to annualised entries+exits. Empty if the workbook has
        no recognisable header or no usable Underground rows.

    Raises
    ------
    ValueError
        If ``data`` is not a readable XLSX workbook.
    """
    import io

    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except (InvalidFileException, KeyError, OSError, zipfile.BadZipFile) as exc:
        raise ValueError("station usage data is not a readable XLSX workbook") from exc

    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    header_idx = _header_row_index(rows)
    if header_idx is None:
        return {}
    header = rows[header_idx]
    group_row = rows[header_idx - 1] if header_idx > 0 else ()
    mode_col = _column_index(header, _USAGE_MODE_COLUMN)
    name_col = _column_index(header, _USAGE_STATION_COLUMN)
    annual_col = _annualised_column_index(header, group_row)
    if mode_col is None or name_col is None or annual_col is None:
        return {}

    usage: dict[str, float] = {}
    for row in rows[header_idx + 1 :]:
        if max(mode_col, name_col, annual_col) >= len(row):
            continue
        mode = row[mode_col]
        if mode is None or str(mode).strip().lower() not in _USAGE_MODES:
            continue
        name = row[name_col]
        annual = _coerce_float(row[annual_col])
        if name is None or annual is None or annual <= 0:
            continue
        for key in usage_match_keys(str(name)):
            if usage.get(key, 0.0) < annual:
                usage[key] = annual
    return usage


def _coerce_float(value: Any) -> float | None:
    """Coerce a spreadsheet cell to a positive ``float`` if possible.

    Parameters
    ----------
    value : Any
        A cell value (number, numeric string, ``"---"`` placeholder, or ``None``).

    Returns
    -------
    float or None
        The float value, or ``None`` if it does not parse as a number.
    """
    if isinstance(value, bool):  # bools are ints in Python; never a count
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(",", "").strip())
        except ValueError:
            return None
    return None


_ABBREVIATIONS: frozenset[str] = frozenset(
    {"st", "mt", "mtn", "ave", "rd", "no", "vol", "etc", "dr", "jr", "sr", "co"}
)
"""Lowercased tokens whose trailing period does not end a sentence (e.g. ``St.``)."""

_MIN_SENTENCE_LEN: int = 25
"""Below this length a period-break is treated as spurious (an abbreviation)."""


def first_sentence(extract: str, max_len: int = 220) -> str | None:
    """Trim a Wikipedia ``extract`` to a single punchy sentence.

    Returns the text up to the first genuine sentence-ending period. A period is
    not treated as a sentence end when it follows a known abbreviation (``St.``,
    ``Mt.``, ...) or when the fragment so far is implausibly short -- this avoids
    truncating "St. John's Wood is a station." to "St.". The result is hard-
    capped at ``max_len``.

    Parameters
    ----------
    extract : str
        The ``extract`` field from a Wikipedia summary.
    max_len : int, optional
        Maximum length of the returned string before truncation.

    Returns
    -------
    str or None
        A trimmed sentence, or ``None`` if the extract is empty.
    """
    text = extract.strip()
    if not text:
        return None

    sentence = text  # fall back to the whole extract if no clean break is found
    for match in re.finditer(r"\.\s", text):
        end = match.start()
        candidate = text[:end]
        # Reject breaks that are too short to be a real sentence...
        if len(candidate) < _MIN_SENTENCE_LEN:
            continue
        # ...or that follow a known abbreviation (the token before the period).
        last_token = re.split(r"[\s(]", candidate)[-1].lower()
        if last_token in _ABBREVIATIONS:
            continue
        sentence = text[: end + 1]
        break

    sentence = sentence.strip()
    if len(sentence) > max_len:
        sentence = sentence[: max_len - 1].rstrip() + "…"
    return sentence or None


# --------------------------------------------------------------------------- #
# Fun-fact selection: skip the dull definition, pick the interesting sentence  #
# --------------------------------------------------------------------------- #

_GENERIC_DEFINITION_RE: re.Pattern[str] = re.compile(
    r"""
    \bis\s+(?:a|an|the)\s+            # "is a" / "is an" / "is the"
    (?:[\w'’\-]+\s+){0,6}?            # up to a few qualifier words (e.g. "Grade II listed")
    (?:
        london\s+underground\s+(?:station|and)  # "London Underground station"/"... and ..."
      | underground\s+station                    # "Underground station"
      | tube\s+station                            # "tube station"
      | (?:railway|train|metro)\s+station\b       # plain "railway/train/metro station"
      | interchange\s+station\b                    # "interchange station"
      | (?:central\s+london\s+)?railway\s+terminus # "central London railway terminus"
    )
    """,
    re.IGNORECASE | re.VERBOSE,
)
"""Matches the dull "X is a[n] ... [Underground] station" opening definition.

Deliberately broad over the boilerplate forms (plain tube station, interchange,
railway terminus) so the heuristic can skip past them to something worth
reading.
"""

_GENERIC_LINE_DEF_RE: re.Pattern[str] = re.compile(
    r"\bis\s+(?:a|an)\s+(?:[\w'’\-]+\s+){0,3}?station\s+on\s+the\b",
    re.IGNORECASE,
)
"""Matches the line-defining opener "X is a station on the <line> line".

Kept separate from :data:`_GENERIC_DEFINITION_RE` because it is pinned to the
indefinite article (``a``/``an``): that distinguishes the definitional "is a
station on the Northern line" from a real superlative like "is the deepest
station on the network", which must NOT be treated as boilerplate.
"""

# Tube line names. A sentence that leans on these is a line spoiler -- the UI
# reveals which lines serve a station only as a post-game reward -- so such
# sentences are demoted, never preferred.
_LINE_NAMES: tuple[str, ...] = (
    "bakerloo",
    "central line",
    "circle line",
    "district line",
    "hammersmith & city",
    "hammersmith and city",
    "jubilee line",
    "metropolitan line",
    "northern line",
    "piccadilly line",
    "victoria line",
    "waterloo & city",
    "waterloo and city",
    "elizabeth line",
)
"""Line-name fragments whose presence makes a sentence a line spoiler."""

_LINE_SPOILER_RE: re.Pattern[str] = re.compile(
    "|".join(re.escape(name) for name in _LINE_NAMES), re.IGNORECASE
)
"""Any tube/Elizabeth line name (case-insensitive), for the spoiler demotion."""

# Phrases that mark a genuinely interesting sentence. Weighted: an explicit
# "named after" is the strongest signal (this is exactly the Oval-cricket-ground
# kind of fact we want), superlatives and notable associations next.
_INTEREST_PATTERNS: tuple[tuple[re.Pattern[str], int], ...] = (
    (re.compile(r"\bnamed (?:after|for|in honou?r of)\b", re.IGNORECASE), 6),
    (re.compile(r"\btakes its name from\b", re.IGNORECASE), 6),
    (re.compile(r"\b(?:named|naming)\b", re.IGNORECASE), 3),
    (
        re.compile(
            r"\b(?:oldest|deepest|highest|busiest|first|only|largest|smallest|"
            r"longest|shortest|newest|rarest|northernmost|southernmost|"
            r"easternmost|westernmost|unique|sole)\b",
            re.IGNORECASE,
        ),
        4,
    ),
    (
        re.compile(
            r"\b(?:film|films|movie|television|tv series|sitcom|novel|song|"
            r"music video|featured in|depicted in|appears in|filmed)\b",
            re.IGNORECASE,
        ),
        4,
    ),
    (
        re.compile(
            r"\b(?:designed by|architect|Charles Holden|listed building|"
            r"grade [I]+ listed|art deco|art-deco|mosaic|murals?|disused|"
            r"abandoned|ghost station|deep-level shelter|air-raid|wartime|"
            r"bomb|World War|blitz)\b",
            re.IGNORECASE,
        ),
        4,
    ),
    (
        re.compile(
            r"\b(?:cricket|football|stadium|ground|palace|cathedral|abbey|"
            r"museum|gallery|prison|gardens?|park|market|cemetery|brewery|"
            r"racecourse|university|hospital)\b",
            re.IGNORECASE,
        ),
        2,
    ),
)
"""Regexes that flag interesting content, paired with the score they add."""

_PROPER_NOUN_RE: re.Pattern[str] = re.compile(r"\b[A-Z][a-z]{2,}\b")
"""A capitalised word (a rough proxy for a named person/place worth mentioning)."""

_SENTENCE_SPLIT_RE: re.Pattern[str] = re.compile(r"(?<=[.!?])\s+(?=[A-Z(])")
"""Splits an intro into sentences at a terminator followed by a capital/paren.

Abbreviations like ``St.`` rarely precede a capital-then-space in these intros,
and the small-fragment guard in :func:`_split_sentences` mops up the rest.
"""


def is_generic_definition(sentence: str) -> bool:
    """Return whether a sentence is the dull "X is a ... station" boilerplate.

    Parameters
    ----------
    sentence : str
        A single sentence.

    Returns
    -------
    bool
        ``True`` if the sentence is the generic Underground-station definition
        (or the line-defining "X is a station on the ... line" form).
    """
    return (
        _GENERIC_DEFINITION_RE.search(sentence) is not None
        or _GENERIC_LINE_DEF_RE.search(sentence) is not None
    )


def _split_sentences(text: str) -> list[str]:
    """Split an intro paragraph into trimmed, non-trivial sentences.

    Fragments shorter than :data:`_MIN_SENTENCE_LEN` are dropped (they are
    almost always split artefacts), and any trailing reference brackets such as
    ``[1]`` are stripped.

    Parameters
    ----------
    text : str
        The intro text (one or more paragraphs).

    Returns
    -------
    list of str
        The cleaned sentences in document order.
    """
    cleaned = re.sub(r"\[\d+\]", "", text).strip()
    if not cleaned:
        return []
    parts = _SENTENCE_SPLIT_RE.split(cleaned)
    out: list[str] = []
    for part in parts:
        candidate = part.strip()
        if len(candidate) >= _MIN_SENTENCE_LEN:
            out.append(candidate)
    return out


def _score_sentence(sentence: str) -> int:
    """Score how interesting a candidate fun-fact sentence is.

    Higher is better. Interest phrases (``named after``, superlatives, films,
    notable landmarks) add their weight; otherwise unseen proper nouns add a
    little (a sentence naming places/people beats a bare locator). A sentence
    that mentions a tube line is heavily penalised so it is never preferred --
    line identities are a post-game reveal in the UI.

    Parameters
    ----------
    sentence : str
        A single candidate sentence.

    Returns
    -------
    int
        The interest score (may be negative for line spoilers).
    """
    score = 0
    for pattern, weight in _INTEREST_PATTERNS:
        if pattern.search(sentence):
            score += weight
    # A couple of proper nouns beyond the leading station name suggest the
    # sentence actually says something specific. Cap the contribution so a
    # name-stuffed locator does not outrank a real fact.
    proper_nouns = _PROPER_NOUN_RE.findall(sentence)
    score += min(max(len(proper_nouns) - 1, 0), 2)
    if _LINE_SPOILER_RE.search(sentence):
        score -= 8
    return score


def best_fun_fact(extract: str, max_len: int = 220) -> str | None:
    """Pick the most interesting sentence from a Wikipedia intro.

    Skips the generic "X is a[n] ... Underground station" definition
    sentence(s) and ranks the remainder by :func:`_score_sentence`, preferring
    "named after", superlatives, notable landmarks/people, and unusual history
    over a bare locator. Ties keep document order (earlier sentences win), which
    favours the lead. Only if nothing non-generic survives does it fall back to
    the first generic definition (then to :func:`first_sentence`), so a fact is
    always returned when the extract has any content.

    Parameters
    ----------
    extract : str
        The Wikipedia intro text (REST summary ``extract`` or the fuller
        Action-API ``prop=extracts`` lead section).
    max_len : int, optional
        Hard cap on the returned sentence length.

    Returns
    -------
    str or None
        One punchy sentence, or ``None`` if the extract is empty.
    """
    sentences = _split_sentences(extract)
    if not sentences:
        return first_sentence(extract, max_len=max_len)

    non_generic = [s for s in sentences if not is_generic_definition(s)]
    if non_generic:
        # argmax by score, keeping the earliest on ties (stable enumerate key).
        best = max(
            enumerate(non_generic),
            key=lambda pair: (_score_sentence(pair[1]), -pair[0]),
        )[1]
    else:
        best = sentences[0]

    best = best.strip()
    if len(best) > max_len:
        best = best[: max_len - 1].rstrip() + "…"
    return best or None


class WikidataClient:
    """Minimal client for the Wikidata SPARQL endpoint.

    Parameters
    ----------
    endpoint : str, optional
        SPARQL endpoint URL. Defaults to :data:`WIKIDATA_SPARQL_URL`.
    timeout : float, optional
        Per-request timeout in seconds. Defaults to :data:`DEFAULT_TIMEOUT`.
    client : httpx.Client or None, optional
        An existing client to use (not closed by this object). When ``None`` a
        client carrying the descriptive User-Agent is created internally.
    """

    def __init__(
        self,
        endpoint: str = WIKIDATA_SPARQL_URL,
        timeout: float = DEFAULT_TIMEOUT,
        client: httpx.Client | None = None,
    ) -> None:
        self.endpoint = endpoint
        self._owns_client = client is None
        self._client = client if client is not None else _new_http_client(timeout)

    def __enter__(self) -> WikidataClient:
        """Enter the runtime context and return this client.

        Returns
        -------
        WikidataClient
            This client instance.
        """
        return self

    def __exit__(self, *exc: object) -> None:
        """Exit the runtime context, closing an internally owned client."""
        self.close()

    def close(self) -> None:
        """Close the underlying HTTP client if this object created it."""
        if self._owns_client:
            self._client.close()

    def fetch_stations(self, query: str = SPARQL_QUERY) -> list[WikidataStation]:
        """Run the SPARQL query and parse station candidates.

        One Wikidata entity can appear in several rows (multiple patronage
        statements, for example); rows are folded so each ``qid`` yields one
        candidate, preferring the largest patronage value seen.

        Parameters
        ----------
        query : str, optional
            SPARQL query string. Defaults to :data:`SPARQL_QUERY`.

        Throttled or transient responses (429/5xx) are retried up to
        :data:`SPARQL_MAX_RETRIES` times, honouring ``Retry-After`` up to
        :data:`SPARQL_MAX_BACKOFF_S`. A persistent error raises so the failure
        is loud rather than silently yielding an empty artefact.

        Returns
        -------
        list of WikidataStation
            All candidates that carried parseable coordinates.

        Raises
        ------
        httpx.HTTPStatusError
            If the SPARQL service keeps returning an error status.
        """
        last_response: httpx.Response | None = None
        for attempt in range(SPARQL_MAX_RETRIES):
            response = self._client.get(
                self.endpoint,
                params={"query": query, "format": "json"},
                headers={"Accept": "application/sparql-results+json"},
            )
            if response.status_code == 200:
                payload: dict[str, Any] = response.json()
                return _parse_sparql_results(payload)
            last_response = response
            if response.status_code not in _RETRYABLE_STATUS:
                break
            if attempt == SPARQL_MAX_RETRIES - 1:
                break
            wait_s = _retry_after_seconds(response, default=2.0 * (attempt + 1))
            if wait_s > SPARQL_MAX_BACKOFF_S:
                # The endpoint wants longer than we will wait; fail loud now.
                break
            time.sleep(wait_s)
        assert last_response is not None  # noqa: S101 - loop always sets it
        last_response.raise_for_status()
        # raise_for_status raises for any non-2xx; this line is unreachable for
        # error statuses but satisfies the type checker for the 2xx-but-unhandled
        # edge (which the loop already returns for).
        payload = last_response.json()  # pragma: no cover
        return _parse_sparql_results(payload)  # pragma: no cover


class WikipediaClient:
    """Minimal client for the Wikipedia REST page-summary endpoint.

    Parameters
    ----------
    base_url : str, optional
        Summary endpoint base. Defaults to :data:`WIKIPEDIA_SUMMARY_URL`.
    query_url : str, optional
        Action API endpoint for the fuller intro. Defaults to
        :data:`WIKIPEDIA_QUERY_URL`.
    timeout : float, optional
        Per-request timeout in seconds. Defaults to ``30.0``.
    delay_s : float, optional
        Seconds to sleep after each request, to stay polite over the ~272
        per-station calls. Defaults to :data:`DEFAULT_WIKI_DELAY_S`. Set to
        ``0`` in tests.
    client : httpx.Client or None, optional
        An existing client to use (not closed by this object). When ``None`` a
        client carrying the descriptive User-Agent is created internally.
    """

    def __init__(
        self,
        base_url: str = WIKIPEDIA_SUMMARY_URL,
        query_url: str = WIKIPEDIA_QUERY_URL,
        timeout: float = 30.0,
        delay_s: float = DEFAULT_WIKI_DELAY_S,
        client: httpx.Client | None = None,
    ) -> None:
        self.base_url = base_url.rstrip("/")
        self.query_url = query_url
        self.delay_s = delay_s
        self._owns_client = client is None
        self._client = client if client is not None else _new_http_client(timeout)

    def __enter__(self) -> WikipediaClient:
        """Enter the runtime context and return this client.

        Returns
        -------
        WikipediaClient
            This client instance.
        """
        return self

    def __exit__(self, *exc: object) -> None:
        """Exit the runtime context, closing an internally owned client."""
        self.close()

    def close(self) -> None:
        """Close the underlying HTTP client if this object created it."""
        if self._owns_client:
            self._client.close()

    def summary_extract(self, title: str) -> str | None:
        """Fetch the plain-text ``extract`` for an article title.

        A missing page (404) or any non-2xx status yields ``None`` rather than
        raising, so a single bad title never aborts the run.

        Parameters
        ----------
        title : str
            Article title, e.g. ``"Oval tube station"``. Spaces are encoded by
            httpx; the endpoint also accepts underscores.

        Returns
        -------
        str or None
            The ``extract`` text, or ``None`` if unavailable.
        """
        from urllib.parse import quote

        url = f"{self.base_url}/{quote(title.replace(' ', '_'), safe='')}"
        try:
            response = self._client.get(url)
        except httpx.HTTPError:
            return None
        finally:
            if self.delay_s > 0:
                time.sleep(self.delay_s)
        if response.status_code != 200:
            return None
        try:
            body: dict[str, Any] = response.json()
        except ValueError:
            return None
        extract = body.get("extract")
        return extract if isinstance(extract, str) and extract.strip() else None

    def intro_extract(self, title: str) -> str | None:
        """Fetch the fuller lead-section plain text via the Action API.

        Uses ``action=query&prop=extracts&exintro&explaintext`` which returns
        the whole intro (often several sentences) rather than the single
        sentence the REST summary can collapse to. Falls back to
        :meth:`summary_extract` on any failure so the caller always has the best
        available text. A missing/empty page yields ``None``.

        Parameters
        ----------
        title : str
            Article title, e.g. ``"Oval tube station"``.

        Returns
        -------
        str or None
            The intro plain text, or ``None`` if unavailable from both
            endpoints.
        """
        params = {
            "action": "query",
            "format": "json",
            "prop": "extracts",
            "exintro": "1",
            "explaintext": "1",
            "redirects": "1",
            "titles": title,
        }
        response: httpx.Response | None
        try:
            response = self._client.get(self.query_url, params=params)
        except httpx.HTTPError:
            response = None
        finally:
            if self.delay_s > 0:
                time.sleep(self.delay_s)
        if response is None or response.status_code != 200:
            return self.summary_extract(title)
        try:
            body: dict[str, Any] = response.json()
        except ValueError:
            return self.summary_extract(title)
        pages = body.get("query", {}).get("pages", {})
        if isinstance(pages, dict):
            for page in pages.values():
                extract = page.get("extract") if isinstance(page, dict) else None
                if isinstance(extract, str) and extract.strip():
                    return extract
        return self.summary_extract(title)


class StationUsageClient:
    """Downloads and parses TfL's Annual Station Counts spreadsheet.

    Fetches the entries+exits workbook once and returns the annualised total per
    Underground station, keyed by graph match key. The single network call lives
    here so tests can drive :func:`parse_station_usage` against an in-memory XLSX
    with no HTTP at all.

    Parameters
    ----------
    url : str, optional
        Workbook URL. Defaults to :data:`STATION_USAGE_URL`.
    timeout : float, optional
        Per-request timeout in seconds. Defaults to :data:`DEFAULT_TIMEOUT`.
    client : httpx.Client or None, optional
        An existing client to use (not closed by this object). When ``None`` a
        client carrying the descriptive User-Agent is created internally.
    """

    def __init__(
        self,
        url: str = STATION_USAGE_URL,
        timeout: float = DEFAULT_TIMEOUT,
        client: httpx.Client | None = None,
    ) -> None:
        self.url = url
        self._owns_client = client is None
        self._client = client if client is not None else _new_http_client(timeout)

    def __enter__(self) -> StationUsageClient:
        """Enter the runtime context and return this client.

        Returns
        -------
        StationUsageClient
            This client instance.
        """
        return self

    def __exit__(self, *exc: object) -> None:
        """Exit the runtime context, closing an internally owned client."""
        self.close()

    def close(self) -> None:
        """Close the underlying HTTP client if this object created it."""
        if self._owns_client:
            self._client.close()

    def fetch_usage(self) -> dict[str, float]:
        """Download the workbook and parse annualised Underground usage.

        A non-2xx response raises so a failed download is loud rather than
        silently yielding empty traffic (the caller can then leave the prior
        artefact intact).

        Returns
        -------
        dict of str to float
            Graph match key to annualised entries+exits.

        Raises
        ------
        httpx.HTTPStatusError
            If the download returns a non-2xx status.
        ValueError
            If the downloaded bytes are not a readable XLSX workbook.
        """
        response = self._client.get(self.url)
        response.raise_for_status()
        return parse_station_usage(response.content)


# --------------------------------------------------------------------------- #
# Optional AI distillation: one grounded fun fact via the Anthropic API        #
# --------------------------------------------------------------------------- #

_AI_SYSTEM_PROMPT: str = (
    "You write a single fun fact about a London Underground station for a trivia "
    "game. Rules, all mandatory:\n"
    "- Use ONLY facts stated in the supplied Wikipedia text. Never add outside "
    "knowledge and never guess. If the text has nothing more interesting than "
    "the station being an Underground station, reply with exactly NONE.\n"
    "- One sentence, at most about 20 words, no preamble.\n"
    "- Do NOT restate that it is a London Underground/tube/railway station.\n"
    "- Do NOT mention which lines serve the station (line names are a spoiler).\n"
    "- Prefer the genuinely interesting angle: what it is named after, a nearby "
    "landmark, a superlative (oldest/deepest/only), an appearance in film or TV, "
    "or unusual history or design.\n"
    "- Output the sentence itself only, or NONE."
)
"""System prompt pinning the model to grounded, spoiler-free, one-line output."""

_AI_REFUSAL: str = "NONE"
"""Sentinel the model is told to return when the text offers no good fact."""


def build_fun_fact_prompt(name: str, intro_text: str) -> str:
    """Build the user message asking the model to distil one fun fact.

    Parameters
    ----------
    name : str
        Human station name (without the ``Underground Station`` suffix is fine).
    intro_text : str
        The Wikipedia intro text the fact must be grounded in.

    Returns
    -------
    str
        The user-turn content for the Messages API.
    """
    trimmed = intro_text.strip()
    return (
        f"Station: {name}\n\n"
        f"Wikipedia text:\n{trimmed}\n\n"
        "Give the single best fun fact following all the rules, or NONE."
    )


def parse_fun_fact_response(payload: dict[str, Any]) -> str | None:
    """Extract the distilled fun fact from a Messages API response body.

    Joins the text blocks of the response, strips wrapping quotes/whitespace,
    and maps the ``NONE`` refusal sentinel (or any line-spoiler / still-generic
    answer) to ``None`` so the caller falls back to the heuristic.

    Parameters
    ----------
    payload : dict
        Decoded Messages API response.

    Returns
    -------
    str or None
        The cleaned one-line fact, or ``None`` if the model declined or returned
        an unusable answer.
    """
    blocks = payload.get("content", [])
    if not isinstance(blocks, list):
        return None
    text = "".join(
        block.get("text", "")
        for block in blocks
        if isinstance(block, dict) and block.get("type") == "text"
    ).strip()
    if not text:
        return None
    # Models sometimes wrap the sentence in quotes; drop a single matching pair.
    if len(text) >= 2 and text[0] in {'"', "'", "“"} and text[-1] in {'"', "'", "”"}:
        text = text[1:-1].strip()
    # Keep only the first line/sentence-ish chunk if it rambled.
    text = text.splitlines()[0].strip()
    if not text or text.strip().rstrip(".").upper() == _AI_REFUSAL:
        return None
    # Guardrails: a line spoiler or a still-generic definition is unusable.
    if _LINE_SPOILER_RE.search(text) or is_generic_definition(text):
        return None
    return text


class AnthropicClient:
    """Minimal Anthropic Messages API client for distilling fun facts.

    Only instantiated when an API key is available. Every failure mode (no key,
    timeout, transport error, non-2xx, malformed body, refusal) resolves to
    ``None`` from :meth:`distil_fun_fact`, so the heuristic always has the last
    word and a flaky API never aborts a regeneration run.

    Parameters
    ----------
    api_key : str
        Anthropic API key (``ANTHROPIC_API_KEY``).
    model : str, optional
        Model id. Defaults to :data:`ANTHROPIC_MODEL` (a Haiku model).
    endpoint : str, optional
        Messages API URL. Defaults to :data:`ANTHROPIC_MESSAGES_URL`.
    timeout : float, optional
        Per-request timeout in seconds. Defaults to :data:`ANTHROPIC_TIMEOUT_S`.
    client : httpx.Client or None, optional
        An existing client to use (not closed by this object). When ``None`` a
        client is created internally.
    """

    def __init__(
        self,
        api_key: str,
        model: str = ANTHROPIC_MODEL,
        endpoint: str = ANTHROPIC_MESSAGES_URL,
        timeout: float = ANTHROPIC_TIMEOUT_S,
        client: httpx.Client | None = None,
    ) -> None:
        self.api_key = api_key
        self.model = model
        self.endpoint = endpoint
        self._owns_client = client is None
        self._client = client if client is not None else httpx.Client(timeout=timeout)

    def __enter__(self) -> AnthropicClient:
        """Enter the runtime context and return this client.

        Returns
        -------
        AnthropicClient
            This client instance.
        """
        return self

    def __exit__(self, *exc: object) -> None:
        """Exit the runtime context, closing an internally owned client."""
        self.close()

    def close(self) -> None:
        """Close the underlying HTTP client if this object created it."""
        if self._owns_client:
            self._client.close()

    def distil_fun_fact(self, name: str, intro_text: str) -> str | None:
        """Ask the model for one grounded, spoiler-free fun fact.

        Parameters
        ----------
        name : str
            Human station name.
        intro_text : str
            Wikipedia intro text to ground the fact in.

        Returns
        -------
        str or None
            The distilled fact, or ``None`` on refusal or any error (the caller
            then falls back to the heuristic).
        """
        if not intro_text.strip():
            return None
        body = {
            "model": self.model,
            "max_tokens": ANTHROPIC_MAX_TOKENS,
            "system": _AI_SYSTEM_PROMPT,
            "messages": [{"role": "user", "content": build_fun_fact_prompt(name, intro_text)}],
        }
        headers = {
            "x-api-key": self.api_key,
            "anthropic-version": ANTHROPIC_VERSION,
            "content-type": "application/json",
        }
        try:
            response = self._client.post(self.endpoint, json=body, headers=headers)
        except httpx.HTTPError:
            return None
        if response.status_code != 200:
            return None
        try:
            payload: dict[str, Any] = response.json()
        except ValueError:
            return None
        return parse_fun_fact_response(payload)


def _new_http_client(timeout: float) -> httpx.Client:
    """Create an httpx client carrying the descriptive User-Agent.

    Parameters
    ----------
    timeout : float
        Per-request timeout in seconds.

    Returns
    -------
    httpx.Client
        A client that follows redirects and sends :data:`USER_AGENT`.
    """
    return httpx.Client(
        timeout=timeout,
        follow_redirects=True,
        headers={"User-Agent": USER_AGENT},
    )


def _retry_after_seconds(response: httpx.Response, default: float) -> float:
    """Read a ``Retry-After`` header as seconds, falling back to a default.

    Only the integer-seconds form is interpreted (the HTTP-date form is rare
    here and treated as the default).

    Parameters
    ----------
    response : httpx.Response
        The throttled response.
    default : float
        Seconds to use when the header is absent or not an integer.

    Returns
    -------
    float
        Seconds to wait before the next attempt.
    """
    raw = response.headers.get("retry-after")
    if raw is None:
        return default
    try:
        return float(int(raw.strip()))
    except (TypeError, ValueError):
        return default


def _parse_sparql_results(payload: dict[str, Any]) -> list[WikidataStation]:
    """Fold raw SPARQL JSON bindings into deduplicated station candidates.

    Parameters
    ----------
    payload : dict
        Decoded ``application/sparql-results+json`` body.

    Returns
    -------
    list of WikidataStation
        One candidate per Wikidata entity with parseable coordinates.
    """
    bindings = payload.get("results", {}).get("bindings", [])
    by_qid: dict[str, WikidataStation] = {}
    for row in bindings:
        coord_cell = row.get("coord", {}).get("value")
        station_uri = row.get("station", {}).get("value")
        if not coord_cell or not station_uri:
            continue
        coords = parse_wkt_point(coord_cell)
        if coords is None:
            continue
        qid = station_uri.rsplit("/", 1)[-1]
        lat, lon = coords

        opened_year: int | None = None
        if "opened" in row:
            opened_year = parse_opened_year(row["opened"].get("value", ""))

        patronage: float | None = None
        if "patronage" in row:
            try:
                patronage = float(row["patronage"].get("value", ""))
            except (TypeError, ValueError):
                patronage = None

        article_title: str | None = None
        if "article" in row:
            article_title = article_title_from_url(row["article"].get("value", ""))

        name = row.get("stationLabel", {}).get("value", qid)

        existing = by_qid.get(qid)
        if existing is None:
            by_qid[qid] = WikidataStation(
                qid=qid,
                name=name,
                lat=lat,
                lon=lon,
                opened_year=opened_year,
                annual_patronage=patronage,
                article_title=article_title,
            )
        else:
            by_qid[qid] = _merge_candidates(existing, opened_year, patronage, article_title)
    return list(by_qid.values())


def _merge_candidates(
    existing: WikidataStation,
    opened_year: int | None,
    patronage: float | None,
    article_title: str | None,
) -> WikidataStation:
    """Fold a repeated SPARQL row into an existing candidate.

    Keeps the earliest opening year, the largest patronage figure, and the
    first article title seen.

    Parameters
    ----------
    existing : WikidataStation
        The candidate accumulated so far.
    opened_year : int or None
        Opening year from the current row.
    patronage : float or None
        Patronage from the current row.
    article_title : str or None
        Article title from the current row.

    Returns
    -------
    WikidataStation
        The merged candidate.
    """
    merged_opened = existing.opened_year
    if opened_year is not None and (merged_opened is None or opened_year < merged_opened):
        merged_opened = opened_year

    merged_patronage = existing.annual_patronage
    if patronage is not None and (merged_patronage is None or patronage > merged_patronage):
        merged_patronage = patronage

    merged_article = existing.article_title or article_title

    return WikidataStation(
        qid=existing.qid,
        name=existing.name,
        lat=existing.lat,
        lon=existing.lon,
        opened_year=merged_opened,
        annual_patronage=merged_patronage,
        article_title=merged_article,
    )


def match_candidate(
    station: GraphStation,
    candidates: list[WikidataStation],
    name_index: dict[str, WikidataStation],
    radius_m: float = DEFAULT_MATCH_RADIUS_M,
) -> WikidataStation | None:
    """Match one graph station to its nearest Wikidata candidate.

    Coordinate proximity is tried first (nearest candidate within ``radius_m``),
    then a normalised-name lookup as a fallback.

    Parameters
    ----------
    station : GraphStation
        The graph station to match.
    candidates : list of WikidataStation
        All Wikidata candidates.
    name_index : dict of str to WikidataStation
        Candidates keyed by :func:`normalise_name` of their label (built once by
        the caller).
    radius_m : float, optional
        Maximum match distance in metres. Defaults to
        :data:`DEFAULT_MATCH_RADIUS_M`.

    Returns
    -------
    WikidataStation or None
        The matched candidate, or ``None`` if neither strategy finds one.
    """
    best: WikidataStation | None = None
    best_dist = radius_m
    for cand in candidates:
        dist = haversine_m(station.lat, station.lon, cand.lat, cand.lon)
        if dist <= best_dist:
            best = cand
            best_dist = dist
    if best is not None:
        return best
    return name_index.get(normalise_name(station.name))


def load_graph_stations(graph_path: str) -> list[GraphStation]:
    """Read the id/name/lat/lon of every station in a ``graph.json``.

    Parameters
    ----------
    graph_path : str
        Path to the graph artefact.

    Returns
    -------
    list of GraphStation
        One entry per station, in file order.
    """
    import json
    from pathlib import Path

    raw = json.loads(Path(graph_path).read_text(encoding="utf-8"))
    return [
        GraphStation(
            id=str(s["id"]),
            name=str(s["name"]),
            lat=float(s["lat"]),
            lon=float(s["lon"]),
        )
        for s in raw.get("stations", [])
    ]


def _assign_ranks(values: dict[str, int], *, descending: bool) -> dict[str, int]:
    """Rank station ids by an integer value, 1 = first.

    Ties share neither value adjustments nor gaps beyond standard competition
    behaviour: ids are ordered by value then id (stable) and assigned a strict
    1..N rank.

    Parameters
    ----------
    values : dict of str to int
        Station id to the value being ranked.
    descending : bool
        ``True`` ranks the largest value first (traffic); ``False`` ranks the
        smallest first (opening year).

    Returns
    -------
    dict of str to int
        Station id to its 1-based rank.
    """
    ordered = sorted(values.items(), key=lambda kv: (-kv[1] if descending else kv[1], kv[0]))
    return {station_id: rank for rank, (station_id, _) in enumerate(ordered, start=1)}


def build_station_infos(
    graph_stations: list[GraphStation],
    candidates: list[WikidataStation],
    wiki: WikipediaClient,
    usage: dict[str, float] | None = None,
    radius_m: float = DEFAULT_MATCH_RADIUS_M,
) -> dict[str, StationInfo]:
    """Match, fetch fun facts, and assemble per-station info (pre-ranking).

    Daily traffic prefers the TfL Annual Station Counts figure (``usage``,
    matched by :func:`normalise_name`); it falls back to the Wikidata patronage
    of the matched candidate only where TfL has no figure for the station.

    Parameters
    ----------
    graph_stations : list of GraphStation
        Stations to enrich.
    candidates : list of WikidataStation
        Wikidata candidates to match against.
    wiki : WikipediaClient
        Client used to fetch Wikipedia summaries for fun facts.
    usage : dict of str to float or None, optional
        Annualised entries+exits keyed by graph match key (from
        :func:`parse_station_usage`). ``None`` disables the TfL source, leaving
        only the sparse Wikidata fallback.
    radius_m : float, optional
        Coordinate match radius in metres.

    Returns
    -------
    dict of str to StationInfo
        One entry per graph station, keyed by graph id. Opened/traffic ranks are
        left unset here and filled by :func:`enrich_stations`.
    """
    usage = usage or {}
    name_index = _build_name_index(candidates)
    infos: dict[str, StationInfo] = {}
    for station in graph_stations:
        match = match_candidate(station, candidates, name_index, radius_m=radius_m)
        opened_year = match.opened_year if match else None
        # Match usage with the same key derivation used to build the dict
        # (usage_match_keys strips TfL mode tails like the ELL/EZL marker), not a
        # plain normalise_name, or stations such as "New Cross ELL" miss.
        annual = next(
            (usage[k] for k in usage_match_keys(station.name) if k in usage),
            None,
        )
        if annual is None and match is not None:
            annual = match.annual_patronage
        daily_traffic = _daily_from_annual(annual)
        wiki_url, fun_fact = _resolve_article(station, match, wiki)
        # Constructed via model_validate with alias keys (matching the Edge /
        # TubeGraph convention in build_graph.py) so the aliased fields are set
        # without alias-named kwargs, keeping mypy happy without the pydantic
        # plugin.
        infos[station.id] = StationInfo.model_validate(
            {
                "name": station.name,
                "openedYear": opened_year,
                "dailyTraffic": daily_traffic,
                "funFact": fun_fact,
                "wikiUrl": wiki_url,
            }
        )
    return infos


def _build_name_index(candidates: list[WikidataStation]) -> dict[str, WikidataStation]:
    """Index candidates by normalised name for fallback matching.

    On a key collision the first candidate wins (Wikidata labels are
    near-unique for tube stations, so collisions are rare).

    Parameters
    ----------
    candidates : list of WikidataStation
        All Wikidata candidates.

    Returns
    -------
    dict of str to WikidataStation
        Normalised label to candidate.
    """
    index: dict[str, WikidataStation] = {}
    for cand in candidates:
        key = normalise_name(cand.name)
        if key and key not in index:
            index[key] = cand
    return index


def _daily_from_annual(annual: float | None) -> int | None:
    """Convert an annual entries+exits figure to a rounded daily figure.

    Parameters
    ----------
    annual : float or None
        Annual passenger usage (TfL annualised count or Wikidata patronage).

    Returns
    -------
    int or None
        ``round(annual / 365)`` when positive, else ``None``.
    """
    if annual is None or annual <= 0:
        return None
    return round(annual / DAYS_PER_YEAR)


def _resolve_article(
    station: GraphStation,
    match: WikidataStation | None,
    wiki: WikipediaClient,
) -> tuple[str | None, str | None]:
    """Resolve the Wikipedia URL and fun fact for a station.

    Uses the matched article title when present, otherwise falls back to
    ``"{name} tube station"`` (with the graph's name stripped of its
    ``Underground Station`` suffix). The URL is only emitted when the summary
    fetch confirms the article exists.

    Parameters
    ----------
    station : GraphStation
        The graph station (used for the fallback title).
    match : WikidataStation or None
        The matched candidate, if any.
    wiki : WikipediaClient
        Client used to fetch the summary.

    Returns
    -------
    tuple of (str or None, str or None)
        ``(wiki_url, fun_fact)``. Either element may be ``None``.
    """
    title = match.article_title if match and match.article_title else None
    if title is None:
        bare = _TITLE_SUFFIX_RE.sub("", station.name).strip()
        title = f"{bare} tube station" if bare else None
    if title is None:
        return None, None

    extract = wiki.summary_extract(title)
    if extract is None:
        return None, None
    from urllib.parse import quote

    wiki_url = f"https://en.wikipedia.org/wiki/{quote(title.replace(' ', '_'), safe='')}"
    return wiki_url, first_sentence(extract)


def enrich_stations(
    graph_stations: list[GraphStation],
    wikidata: WikidataClient,
    wikipedia: WikipediaClient,
    generated_at: str,
    usage_client: StationUsageClient | None = None,
    radius_m: float = DEFAULT_MATCH_RADIUS_M,
) -> StationInfoFile:
    """Produce the full station-info artefact for a set of graph stations.

    Fetches Wikidata candidates and the TfL station-usage figures once, matches
    and enriches each station, then computes network-wide ranks (``openedRank``
    1 = oldest; ``dailyTrafficRank`` 1 = busiest) among the stations that have
    each stat. Daily traffic comes from TfL where available, falling back to
    Wikidata patronage; the rank is computed over whatever coverage results.

    Parameters
    ----------
    graph_stations : list of GraphStation
        Stations to enrich (typically every station in ``graph.json``).
    wikidata : WikidataClient
        Client used for the single SPARQL query.
    wikipedia : WikipediaClient
        Client used for per-station summary fetches.
    generated_at : str
        ISO date written to the artefact's ``generatedAt`` field.
    usage_client : StationUsageClient or None, optional
        Client for TfL's Annual Station Counts. When ``None``, daily traffic
        relies solely on the sparse Wikidata patronage fallback. The download
        runs here so any failure propagates to the caller (loud, not silent).
    radius_m : float, optional
        Coordinate match radius in metres.

    Returns
    -------
    StationInfoFile
        The assembled, ranked artefact.

    Raises
    ------
    httpx.HTTPStatusError
        If ``usage_client`` is given and its download returns a non-2xx status.
    ValueError
        If ``usage_client`` is given and the downloaded bytes are not a readable
        XLSX workbook.
    """
    candidates = wikidata.fetch_stations()
    usage = usage_client.fetch_usage() if usage_client is not None else {}
    infos = build_station_infos(
        graph_stations, candidates, wikipedia, usage=usage, radius_m=radius_m
    )

    opened_values = {
        sid: info.opened_year for sid, info in infos.items() if info.opened_year is not None
    }
    traffic_values = {
        sid: info.daily_traffic for sid, info in infos.items() if info.daily_traffic is not None
    }
    opened_ranks = _assign_ranks(opened_values, descending=False)
    traffic_ranks = _assign_ranks(traffic_values, descending=True)

    for sid, info in infos.items():
        if sid in opened_ranks:
            info.opened_rank = opened_ranks[sid]
        if sid in traffic_ranks:
            info.daily_traffic_rank = traffic_ranks[sid]

    # Constructed via model_validate with alias keys, matching the Edge /
    # TubeGraph convention in build_graph.py (the pydantic mypy plugin is not
    # enabled, so alias-named kwargs would not type-check).
    counts = StationInfoCounts.model_validate(
        {
            "total": len(graph_stations),
            "withOpened": len(opened_values),
            "withTraffic": len(traffic_values),
        }
    )
    return StationInfoFile.model_validate(
        {
            "version": "1.0",
            "generatedAt": generated_at,
            "counts": counts,
            "stations": infos,
        }
    )


def clean_station_name(name: str) -> str:
    """Strip the ``Underground Station`` suffix and parentheticals from a name.

    Turns ``"Oval Underground Station"`` into ``"Oval"`` and
    ``"Paddington (H&C Line)-Underground"`` into ``"Paddington"`` -- the form
    wanted for an AI prompt or a fallback article title. An internal ``Station``
    (``Battersea Power Station``) is preserved.

    Parameters
    ----------
    name : str
        Raw display name from the artefact.

    Returns
    -------
    str
        The cleaned, human-readable station name.
    """
    # Drop any parenthetical disambiguator anywhere (e.g. "(H&C Line)"), then the
    # trailing "[-/ ]Underground[ Station]" / "tube station" suffix. Done in this
    # order because the suffix strip in _TITLE_SUFFIX_RE is end-anchored and would
    # otherwise leave a parenthetical that is not flush with the end of the name.
    cleaned = re.sub(r"\s*\([^)]*\)", "", name)
    # Strip repeatedly so stacked suffixes collapse, e.g.
    # "New Cross ELL Rail Station" -> "New Cross ELL Rail" -> "New Cross".
    while True:
        stripped = _TITLE_SUFFIX_RE.sub("", cleaned).strip()
        if stripped == cleaned:
            return stripped
        cleaned = stripped


def select_fun_fact(
    name: str,
    intro_text: str | None,
    ai_client: AnthropicClient | None = None,
) -> str | None:
    """Choose one fun fact for a station from its Wikipedia intro.

    Prefers the AI distillation when an :class:`AnthropicClient` is supplied and
    returns a usable answer; otherwise (or on AI refusal/error) falls back to the
    :func:`best_fun_fact` heuristic. Returns ``None`` only when there is no intro
    text at all.

    Parameters
    ----------
    name : str
        Cleaned station name (see :func:`clean_station_name`).
    intro_text : str or None
        The fetched Wikipedia intro text, or ``None`` if the fetch failed.
    ai_client : AnthropicClient or None, optional
        When provided, tried first for a distilled fact.

    Returns
    -------
    str or None
        The chosen fun fact, or ``None`` if no intro text was available.
    """
    if not intro_text or not intro_text.strip():
        return None
    if ai_client is not None:
        distilled = ai_client.distil_fun_fact(name, intro_text)
        if distilled:
            return distilled
    return best_fun_fact(intro_text)


def _looks_like_line_article(title: str | None) -> bool:
    """Return whether an article title is a tube-line article, not a station.

    A handful of artefact ``wikiUrl``s point at a line article (e.g. Aldgate ->
    ``Metropolitan_line``) rather than the station, a legacy of an imperfect
    Wikidata sitelink. Such an intro is all line spoiler, so the caller refetches
    from the canonical ``"{name} tube station"`` title instead.

    Parameters
    ----------
    title : str or None
        Article title decoded from the ``wikiUrl``.

    Returns
    -------
    bool
        ``True`` if the title is (or ends with) a tube-line article title.
    """
    if not title:
        return False
    return bool(re.search(r"\b(?:line|bakerloo)\s*$", title, re.IGNORECASE))


def _fetch_station_fact(
    name: str,
    wiki_url: str | None,
    wiki: WikipediaClient,
    ai_client: AnthropicClient | None,
) -> str | None:
    """Fetch the intro for one station and distil its fun fact.

    The article title is recovered from the existing ``wiki_url`` so the fact
    stays grounded in the very page the artefact links. If that title is missing
    or is a line article (a known legacy quirk), the canonical
    ``"{clean name} tube station"`` title is used so the fact describes the
    station, not a line. The ``wiki_url`` itself is never changed by this
    function -- only the fact is produced.

    Parameters
    ----------
    name : str
        Cleaned station name.
    wiki_url : str or None
        The station's existing Wikipedia URL.
    wiki : WikipediaClient
        Client for the intro fetch.
    ai_client : AnthropicClient or None
        Optional AI distiller, tried before the heuristic.

    Returns
    -------
    str or None
        The chosen fun fact, or ``None`` if no usable intro was found.
    """
    title = article_title_from_url(wiki_url) if wiki_url else None
    if title is not None and not _looks_like_line_article(title):
        intro = wiki.intro_extract(title)
        fact = select_fun_fact(name, intro, ai_client)
        if fact and not is_generic_definition(fact):
            return fact
        # Hold the generic fact as a last resort but try the canonical title for
        # something better.
        fallback_fact = fact
    else:
        fallback_fact = None

    canonical = f"{name} tube station" if name else None
    if canonical and canonical != title:
        intro = wiki.intro_extract(canonical)
        fact = select_fun_fact(name, intro, ai_client)
        if fact:
            return fact
    return fallback_fact


def refresh_fun_facts(
    info_file: StationInfoFile,
    wiki: WikipediaClient,
    ai_client: AnthropicClient | None = None,
) -> StationInfoFile:
    """Return a copy of the artefact with only each station's ``funFact`` redone.

    For every station the Wikipedia intro is fetched (from the station's existing
    ``wikiUrl``, see :func:`_fetch_station_fact`) and distilled into one punchy,
    sourced, spoiler-free sentence -- via the AI path when ``ai_client`` is given,
    otherwise the :func:`best_fun_fact` heuristic. Every other field (``name``,
    ``openedYear``, ``openedRank``, ``dailyTraffic``, ``dailyTrafficRank``,
    ``wikiUrl``) and the file-level ``version``/``generatedAt``/``counts`` are
    copied through unchanged. A station whose fetch yields nothing keeps its
    current fact rather than losing it.

    Parameters
    ----------
    info_file : StationInfoFile
        The existing, validated artefact to refresh.
    wiki : WikipediaClient
        Client for the per-station intro fetches.
    ai_client : AnthropicClient or None, optional
        Optional AI distiller.

    Returns
    -------
    StationInfoFile
        A new, re-validated artefact identical to the input except for refreshed
        ``funFact`` values.
    """
    # Curated overrides (hand-authored, authoritative) keyed by normalised name.
    curated_by_norm = {normalise_name(k): v for k, v in CURATED_FACTS.items()}

    stations: dict[str, dict[str, Any]] = {}
    for sid, info in info_file.stations.items():
        data = info.model_dump(by_alias=True, exclude_none=True)
        name = clean_station_name(info.name)
        # Match on the cleaned name so the override keys (cleaned display names)
        # and the station names go through the same suffix stripping (e.g. the
        # internal "ELL" marker), which plain normalise_name does not remove.
        curated = curated_by_norm.get(normalise_name(name))
        if curated:
            # A curated fact wins, and we skip the (slow) Wikipedia fetch entirely.
            data["funFact"] = curated
        else:
            fact = _fetch_station_fact(name, info.wiki_url, wiki, ai_client)
            if fact:
                data["funFact"] = fact
            # else: leave the pre-existing funFact (if any) untouched.
        stations[sid] = data

    return StationInfoFile.model_validate(
        {
            "version": info_file.version,
            "generatedAt": info_file.generated_at,
            "counts": info_file.counts.model_dump(by_alias=True),
            "stations": stations,
        }
    )


def apply_curated_stats(info_file: StationInfoFile) -> StationInfoFile:
    """Fill missing stats from the curated overrides and recompute the ranks.

    For every station whose ``openedYear`` or ``dailyTraffic`` is absent, the
    curated override (``curated_stats``, matched on the cleaned name like the
    fun-fact overrides) is applied; values already resolved from an automated
    source are never overwritten -- except by the explicit
    ``CURATED_YEAR_CORRECTIONS``, which fix known-bad Wikidata opening dates
    unconditionally. Both ``openedRank`` and ``dailyTrafficRank``
    are then recomputed across the full post-fill coverage -- the ranks are
    relative claims ("Nth oldest"), so they must always be derived from the
    whole population, never hand-patched per station. The file-level ``counts``
    are refreshed to match. Pure merge: no network access.

    Parameters
    ----------
    info_file : StationInfoFile
        The existing, validated artefact.

    Returns
    -------
    StationInfoFile
        A new, re-validated artefact with gaps filled and ranks recomputed.
    """
    opened_by_norm = {normalise_name(k): v for k, v in CURATED_OPENED_YEARS.items()}
    traffic_by_norm = {normalise_name(k): v for k, v in CURATED_DAILY_TRAFFIC.items()}
    corrections_by_norm = {normalise_name(k): v for k, v in CURATED_YEAR_CORRECTIONS.items()}

    stations: dict[str, dict[str, Any]] = {}
    opened_values: dict[str, int] = {}
    traffic_values: dict[str, int] = {}
    for sid, info in info_file.stations.items():
        data = info.model_dump(by_alias=True, exclude_none=True)
        key = normalise_name(clean_station_name(info.name))
        if key in corrections_by_norm:
            data["openedYear"] = corrections_by_norm[key]
        elif info.opened_year is None and key in opened_by_norm:
            data["openedYear"] = opened_by_norm[key]
        if info.daily_traffic is None and key in traffic_by_norm:
            data["dailyTraffic"] = traffic_by_norm[key]
        if "openedYear" in data:
            opened_values[sid] = int(data["openedYear"])
        if "dailyTraffic" in data:
            traffic_values[sid] = int(data["dailyTraffic"])
        stations[sid] = data

    opened_ranks = _assign_ranks(opened_values, descending=False)
    traffic_ranks = _assign_ranks(traffic_values, descending=True)
    for sid, data in stations.items():
        data.pop("openedRank", None)
        data.pop("dailyTrafficRank", None)
        if sid in opened_ranks:
            data["openedRank"] = opened_ranks[sid]
        if sid in traffic_ranks:
            data["dailyTrafficRank"] = traffic_ranks[sid]

    return StationInfoFile.model_validate(
        {
            "version": info_file.version,
            "generatedAt": info_file.generated_at,
            "counts": {
                "total": len(stations),
                "withOpened": len(opened_values),
                "withTraffic": len(traffic_values),
            },
            "stations": stations,
        }
    )


def load_station_info_file(path: str) -> StationInfoFile:
    """Load and validate an existing ``stations-info.json`` artefact.

    Parameters
    ----------
    path : str
        Path to the artefact.

    Returns
    -------
    StationInfoFile
        The parsed, validated artefact (camelCase keys are accepted via the
        model's field aliases).
    """
    import json
    from pathlib import Path

    raw = json.loads(Path(path).read_text(encoding="utf-8"))
    return StationInfoFile.model_validate(raw)


def write_station_info(info_file: StationInfoFile, out_path: str) -> None:
    """Serialise a station-info artefact to JSON on disk.

    Emits camelCase keys (``by_alias=True``) and omits absent optional fields
    (``exclude_none=True``), pretty-printed with two-space indentation and a
    trailing newline. Parent directories are created if missing.

    Parameters
    ----------
    info_file : StationInfoFile
        The artefact to serialise.
    out_path : str
        Destination file path.
    """
    import json
    from pathlib import Path

    destination = Path(out_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    payload = info_file.model_dump(by_alias=True, exclude_none=True)
    destination.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
